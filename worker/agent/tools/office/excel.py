"""Excel/Spreadsheet tool - create, read, and manipulate spreadsheets."""

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent.tools.base import BaseTool


class ExcelTool(BaseTool):
    name = "excel"
    description = (
        "Create, read, and manipulate Excel spreadsheets (.xlsx). "
        "Supports: creating workbooks, adding data, formulas, charts, formatting, "
        "pivot-like summaries, multiple sheets, and conditional formatting."
    )
    parameters = {
        "type": "object",
        "properties": {
            "action": {
                "type": "string",
                "enum": [
                    "create",
                    "read",
                    "add_data",
                    "add_formula",
                    "add_chart",
                    "format_cells",
                    "add_sheet",
                    "auto_width",
                ],
                "description": "Action to perform",
            },
            "filepath": {
                "type": "string",
                "description": "Path to the Excel file",
            },
            "sheet_name": {
                "type": "string",
                "description": "Name of the sheet (default: 'Sheet')",
            },
            "data": {
                "type": "array",
                "description": "2D array of data (rows of cells)",
                "items": {"type": "array", "items": {}},
            },
            "start_cell": {
                "type": "string",
                "description": "Starting cell (e.g., 'A1')",
            },
            "formula": {
                "type": "string",
                "description": "Excel formula (e.g., '=SUM(A1:A10)')",
            },
            "chart_type": {
                "type": "string",
                "enum": ["bar", "line", "pie"],
                "description": "Type of chart",
            },
            "chart_range": {
                "type": "string",
                "description": "Data range for chart (e.g., 'A1:B10')",
            },
            "format_options": {
                "type": "object",
                "description": "Formatting options (bold, color, fill, etc.)",
            },
        },
        "required": ["action", "filepath"],
    }

    async def execute(self, **kwargs: Any) -> Any:
        action = kwargs["action"]
        filepath = kwargs["filepath"]

        if action == "create":
            return self._create(filepath, kwargs.get("sheet_name", "Sheet"))
        elif action == "read":
            return self._read(filepath, kwargs.get("sheet_name"))
        elif action == "add_data":
            return self._add_data(
                filepath,
                kwargs.get("data", []),
                kwargs.get("sheet_name"),
                kwargs.get("start_cell", "A1"),
            )
        elif action == "add_formula":
            return self._add_formula(
                filepath,
                kwargs["start_cell"],
                kwargs["formula"],
                kwargs.get("sheet_name"),
            )
        elif action == "add_chart":
            return self._add_chart(
                filepath,
                kwargs.get("chart_type", "bar"),
                kwargs.get("chart_range", ""),
                kwargs.get("sheet_name"),
            )
        elif action == "format_cells":
            return self._format_cells(
                filepath,
                kwargs.get("start_cell", "A1"),
                kwargs.get("format_options", {}),
                kwargs.get("sheet_name"),
            )
        elif action == "add_sheet":
            return self._add_sheet(filepath, kwargs.get("sheet_name", "Sheet2"))
        elif action == "auto_width":
            return self._auto_width(filepath, kwargs.get("sheet_name"))

        return f"Unknown action: {action}"

    def _create(self, filepath: str, sheet_name: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        return f"Created Excel file: {filepath}"

    def _read(self, filepath: str, sheet_name: str | None) -> str:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])
        return json.dumps({"sheets": wb.sheetnames, "data": data, "rows": len(data)})

    def _add_data(
        self, filepath: str, data: list, sheet_name: str | None, start_cell: str
    ) -> str:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        start_row = int("".join(c for c in start_cell if c.isdigit()))
        start_col = sum(
            (ord(c.upper()) - ord("A") + 1)
            for c in start_cell
            if c.isalpha()
        )

        for r_idx, row in enumerate(data):
            for c_idx, value in enumerate(row):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

        wb.save(filepath)
        return f"Added {len(data)} rows of data starting at {start_cell}"

    def _add_formula(
        self, filepath: str, cell: str, formula: str, sheet_name: str | None
    ) -> str:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws[cell] = formula
        wb.save(filepath)
        return f"Added formula {formula} to cell {cell}"

    def _add_chart(
        self, filepath: str, chart_type: str, chart_range: str, sheet_name: str | None
    ) -> str:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active

        chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        chart_cls = chart_classes.get(chart_type, BarChart)
        chart = chart_cls()

        if chart_range:
            parts = chart_range.split(":")
            if len(parts) == 2:
                data_ref = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)

        ws.add_chart(chart, "E2")
        wb.save(filepath)
        return f"Added {chart_type} chart to the sheet"

    def _format_cells(
        self, filepath: str, cell_range: str, options: dict, sheet_name: str | None
    ) -> str:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active

        font_kwargs = {}
        if options.get("bold"):
            font_kwargs["bold"] = True
        if options.get("font_size"):
            font_kwargs["size"] = options["font_size"]
        if options.get("font_color"):
            font_kwargs["color"] = options["font_color"]

        for row in ws[cell_range]:
            for cell in row if isinstance(row, tuple) else [row]:
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
                if options.get("fill_color"):
                    cell.fill = PatternFill(
                        start_color=options["fill_color"],
                        end_color=options["fill_color"],
                        fill_type="solid",
                    )
                if options.get("align"):
                    cell.alignment = Alignment(horizontal=options["align"])

        wb.save(filepath)
        return f"Formatted cells {cell_range}"

    def _add_sheet(self, filepath: str, sheet_name: str) -> str:
        wb = openpyxl.load_workbook(filepath)
        wb.create_sheet(sheet_name)
        wb.save(filepath)
        return f"Added sheet: {sheet_name}"

    def _auto_width(self, filepath: str, sheet_name: str | None) -> str:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(filepath)
        return "Auto-adjusted column widths"
